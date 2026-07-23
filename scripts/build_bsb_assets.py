#!/usr/bin/env python3
"""Build local, chapter-sized Berean Standard Bible reader assets.

This importer is deliberately separate from the minutes build.  It accepts a
versioned BSB USFM ZIP (or the publisher's spreadsheet) downloaded beforehand,
writes deterministic chapter JSON files, and records the exact source digest
in the repository metadata.

Usage:
  python3 scripts/build_bsb_assets.py /path/to/bsb_usfm.zip
  python3 scripts/build_bsb_assets.py --check
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
SCRIPTURE_DIR = ROOT / "scripture"
ASSET_DIR = ROOT / "assets" / "scripture" / "bsb"
METADATA_PATH = SCRIPTURE_DIR / "bible-books.json"
PROVENANCE_PATH = SCRIPTURE_DIR / "bsb-source.json"

# The project owns this canonical ordering and matching vocabulary.  Chapter
# and verse counts are emitted from the pinned BSB source during import, then
# checked below; they are not inferred by the citation parser at runtime.
CANON: list[tuple[str, str, list[str], bool]] = [
    ("Gen", "Genesis", ["Gen", "Ge"], False),
    ("Exod", "Exodus", ["Exod", "Ex"], False),
    ("Lev", "Leviticus", ["Lev"], False),
    ("Num", "Numbers", ["Num", "Numb"], False),
    ("Deut", "Deuteronomy", ["Deut", "Dt"], False),
    ("Josh", "Joshua", ["Josh"], False),
    ("Judg", "Judges", ["Judg"], False),
    ("Ruth", "Ruth", [], False),
    ("1Sam", "1 Samuel", ["1 Sam", "1 Sa", "I Sam", "I Sa"], False),
    ("2Sam", "2 Samuel", ["2 Sam", "2 Sa", "II Sam", "II Sa"], False),
    ("1Kgs", "1 Kings", ["1 Kings", "1 Kgs", "I Kings", "I Kgs"], False),
    ("2Kgs", "2 Kings", ["2 Kings", "2 Kgs", "II Kings", "II Kgs"], False),
    ("1Chr", "1 Chronicles", ["1 Chron", "1 Chr", "I Chron", "I Chr"], False),
    ("2Chr", "2 Chronicles", ["2 Chron", "2 Chr", "II Chron", "II Chr"], False),
    ("Ezra", "Ezra", [], False),
    ("Neh", "Nehemiah", ["Neh"], False),
    ("Esth", "Esther", ["Esth"], False),
    ("Job", "Job", [], False),
    ("Ps", "Psalms", ["Psalm", "Ps", "Psa", "Pss"], False),
    ("Prov", "Proverbs", ["Prov", "Pr"], False),
    ("Eccl", "Ecclesiastes", ["Eccl", "Ecc"], False),
    ("Song", "Song of Solomon", ["Song of Songs", "Song"], False),
    ("Isa", "Isaiah", ["Isa", "Is"], False),
    ("Jer", "Jeremiah", ["Jer"], False),
    ("Lam", "Lamentations", ["Lam"], False),
    ("Ezek", "Ezekiel", ["Ezek", "Eze"], False),
    ("Dan", "Daniel", ["Dan", "Da"], False),
    ("Hos", "Hosea", ["Hos", "Ho"], False),
    ("Joel", "Joel", [], False),
    ("Amos", "Amos", ["Am"], False),
    ("Obad", "Obadiah", ["Obad", "Ob"], True),
    ("Jonah", "Jonah", ["Jon"], False),
    ("Mic", "Micah", ["Mic"], False),
    ("Nah", "Nahum", ["Nah", "Na"], False),
    ("Hab", "Habakkuk", ["Hab", "Hb"], False),
    ("Zeph", "Zephaniah", ["Zeph", "Zep"], False),
    ("Hag", "Haggai", ["Hag", "Hg"], False),
    ("Zech", "Zechariah", ["Zech", "Zec"], False),
    ("Mal", "Malachi", ["Mal"], False),
    ("Matt", "Matthew", ["Matt", "Mt"], False),
    ("Mark", "Mark", ["Mk"], False),
    ("Luke", "Luke", ["Lk"], False),
    ("John", "John", ["Jn"], False),
    ("Acts", "Acts", ["Ac"], False),
    ("Rom", "Romans", ["Rom", "Ro"], False),
    ("1Cor", "1 Corinthians", ["1 Cor", "1 Co", "I Cor", "I Co"], False),
    ("2Cor", "2 Corinthians", ["2 Cor", "2 Co", "II Cor", "II Co"], False),
    ("Gal", "Galatians", ["Gal"], False),
    ("Eph", "Ephesians", ["Eph"], False),
    ("Phil", "Philippians", ["Phil", "Php"], False),
    ("Col", "Colossians", ["Col"], False),
    ("1Thess", "1 Thessalonians", ["1 Thess", "1 Th", "I Thess", "I Th"], False),
    ("2Thess", "2 Thessalonians", ["2 Thess", "2 Th", "II Thess", "II Th"], False),
    ("1Tim", "1 Timothy", ["1 Tim", "1 Ti", "I Tim", "I Ti"], False),
    ("2Tim", "2 Timothy", ["2 Tim", "2 Ti", "II Tim", "II Ti"], False),
    ("Titus", "Titus", ["Tit"], False),
    ("Phlm", "Philemon", ["Phlm", "Phm"], True),
    ("Heb", "Hebrews", ["Heb"], False),
    ("Jas", "James", ["Jas", "Jm"], False),
    ("1Pet", "1 Peter", ["1 Pet", "1 Pe", "I Pet", "I Pe"], False),
    ("2Pet", "2 Peter", ["2 Pet", "2 Pe", "II Pet", "II Pe"], False),
    ("1John", "1 John", ["1 Jn", "I John", "I Jn"], False),
    ("2John", "2 John", ["2 Jn", "II John", "II Jn"], True),
    ("3John", "3 John", ["3 Jn", "III John", "III Jn"], True),
    ("Jude", "Jude", [], True),
    ("Rev", "Revelation", ["Rev", "Re"], False),
]

BOOK_BY_SOURCE_NAME = {
    "Song of Solomon": "Song",
    "Song of Songs": "Song",
    "Psalm": "Ps",
    **{name: ident for ident, name, _aliases, _one_chapter in CANON},
}

BOOK_BY_USFM_ID = {
    "GEN": "Gen", "EXO": "Exod", "LEV": "Lev", "NUM": "Num", "DEU": "Deut", "JOS": "Josh",
    "JDG": "Judg", "RUT": "Ruth", "1SA": "1Sam", "2SA": "2Sam", "1KI": "1Kgs", "2KI": "2Kgs",
    "1CH": "1Chr", "2CH": "2Chr", "EZR": "Ezra", "NEH": "Neh", "EST": "Esth", "JOB": "Job",
    "PSA": "Ps", "PRO": "Prov", "ECC": "Eccl", "SNG": "Song", "ISA": "Isa", "JER": "Jer",
    "LAM": "Lam", "EZK": "Ezek", "DAN": "Dan", "HOS": "Hos", "JOL": "Joel", "AMO": "Amos",
    "OBA": "Obad", "JON": "Jonah", "MIC": "Mic", "NAM": "Nah", "HAB": "Hab", "ZEP": "Zeph",
    "HAG": "Hag", "ZEC": "Zech", "MAL": "Mal", "MAT": "Matt", "MRK": "Mark", "LUK": "Luke",
    "JHN": "John", "ACT": "Acts", "ROM": "Rom", "1CO": "1Cor", "2CO": "2Cor", "GAL": "Gal",
    "EPH": "Eph", "PHP": "Phil", "COL": "Col", "1TH": "1Thess", "2TH": "2Thess", "1TI": "1Tim",
    "2TI": "2Tim", "TIT": "Titus", "PHM": "Phlm", "HEB": "Heb", "JAS": "Jas", "1PE": "1Pet",
    "2PE": "2Pet", "1JN": "1John", "2JN": "2John", "3JN": "3John", "JUD": "Jude", "REV": "Rev",
}


def clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def read_bsb_spreadsheet(source: Path) -> dict[str, dict[int, dict[int, str]]]:
    """Collect BSB word-table rows into ordinary verse strings."""
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook["biblosinterlinear96"]
    verses: dict[str, dict[int, dict[int, list[str]]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(list))
    )
    current: tuple[str, int, int] | None = None

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row_number % 50_000 == 0:
            print(f"Read {row_number:,} BSB source rows…", flush=True)
        if len(row) < 20 or not isinstance(row[6], str):
            continue
        ref = row[12]
        if isinstance(ref, str):
            match = re.fullmatch(r"(.+?) (\d+):(\d+)", ref.strip())
            if not match:
                continue
            source_name, chapter, verse = match.groups()
            book = BOOK_BY_SOURCE_NAME.get(source_name)
            if not book:
                raise ValueError(f"Unmapped BSB book name: {source_name}")
            current = (book, int(chapter), int(verse))
        if not current:
            continue
        # Columns R through W are the text fragments used by the upstream
        # normalizer.  Keep the BSB wording but omit interlinear apparatus.
        fragments = [item for item in row[17:23] if isinstance(item, str)]
        if fragments:
            verses[current[0]][current[1]][current[2]].append("".join(fragments))

    return {
        book: {
            chapter: {verse: clean_text("".join(parts)) for verse, parts in chapter_data.items()}
            for chapter, chapter_data in book_data.items()
        }
        for book, book_data in verses.items()
    }


def usfm_text(value: str) -> str:
    """Keep readable Scripture wording while discarding USFM structure/notes."""
    value = re.sub(r"\\f\b.*?\\f\*", "", value, flags=re.DOTALL)
    value = re.sub(r"\\x\b.*?\\x\*", "", value, flags=re.DOTALL)
    value = re.sub(r"\\[A-Za-z0-9]+\*?", "", value)
    return clean_text(value)


def read_bsb_usfm(source: Path) -> dict[str, dict[int, dict[int, str]]]:
    """Read the publisher's compact BSB USFM ZIP without an external parser."""
    if source.is_dir():
        files = sorted(source.glob("*.usfm"))
        contents = [(path.name, path.read_text(encoding="utf-8")) for path in files]
    else:
        with zipfile.ZipFile(source) as archive:
            contents = [
                (name.rsplit("/", 1)[-1], archive.read(name).decode("utf-8"))
                for name in sorted(archive.namelist()) if name.lower().endswith(".usfm")
            ]

    collected: dict[str, dict[int, dict[int, list[str]]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(list))
    )
    for filename, source_text in contents:
        book_match = re.search(r"^\\id\s+([A-Z0-9]+)\b", source_text, re.MULTILINE)
        if not book_match:
            raise ValueError(f"Missing USFM book ID in {filename}")
        book = BOOK_BY_USFM_ID.get(book_match.group(1))
        if not book:
            raise ValueError(f"Unmapped USFM book ID {book_match.group(1)} in {filename}")
        chapter = 0
        current_verse: int | None = None
        for raw_line in source_text.splitlines():
            chapter_match = re.match(r"^\\c\s+(\d+)\b", raw_line)
            if chapter_match:
                chapter = int(chapter_match.group(1))
                current_verse = None
                continue
            if re.match(r"^\\(?:id|usfm|h|toc\d*|mt\d*|s\d*|r)\b", raw_line):
                continue
            markers = list(re.finditer(r"\\v\s+(\d+)\s*", raw_line))
            if not markers:
                if chapter and current_verse:
                    text = usfm_text(raw_line)
                    if text:
                        collected[book][chapter][current_verse].append(text)
                continue
            if current_verse and markers[0].start() > 0:
                prefix = usfm_text(raw_line[:markers[0].start()])
                if prefix:
                    collected[book][chapter][current_verse].append(prefix)
            for index, marker in enumerate(markers):
                if not chapter:
                    raise ValueError(f"Verse appears before a chapter in {filename}")
                current_verse = int(marker.group(1))
                end = markers[index + 1].start() if index + 1 < len(markers) else len(raw_line)
                text = usfm_text(raw_line[marker.end():end])
                if text:
                    collected[book][chapter][current_verse].append(text)
    return {
        book: {
            chapter: {verse: clean_text(" ".join(parts)) for verse, parts in chapter_data.items()}
            for chapter, chapter_data in book_data.items()
        }
        for book, book_data in collected.items()
    }


def read_bsb(source: Path) -> dict[str, dict[int, dict[int, str]]]:
    return read_bsb_usfm(source) if source.suffix.lower() == ".zip" or source.is_dir() else read_bsb_spreadsheet(source)


def build(source: Path) -> None:
    source_sha = hashlib.sha256(source.read_bytes()).hexdigest()
    print("Reading the pinned BSB source…", flush=True)
    verses = read_bsb(source)
    if set(verses) != {book_id for book_id, _name, _aliases, _single in CANON}:
        missing = {book_id for book_id, _name, _aliases, _single in CANON} - set(verses)
        extra = set(verses) - {book_id for book_id, _name, _aliases, _single in CANON}
        raise ValueError(f"BSB source does not match the 66-book canon; missing={missing}, extra={extra}")

    SCRIPTURE_DIR.mkdir(parents=True, exist_ok=True)
    if ASSET_DIR.exists():
        shutil.rmtree(ASSET_DIR)
    ASSET_DIR.mkdir(parents=True)

    books: list[dict[str, Any]] = []
    for order, (book_id, name, aliases, one_chapter) in enumerate(CANON, start=1):
        print(f"Writing {name}…", flush=True)
        book_data = verses[book_id]
        chapters = sorted(book_data)
        if chapters != list(range(1, len(chapters) + 1)):
            raise ValueError(f"{name} has non-contiguous chapters in the BSB source")
        verse_counts: list[int] = []
        for chapter in chapters:
            chapter_data = book_data[chapter]
            verse_numbers = sorted(chapter_data)
            # A modern critical-text translation can intentionally omit a
            # traditional verse number.  Keep the conventional upper bound
            # for fast bounds checks and the exact numbered-verse list for
            # display validation; do not invent text for an absent verse.
            verse_counts.append(max(verse_numbers))
            payload = {
                "version": 1,
                "translation": "Berean Standard Bible (BSB)",
                "book": book_id,
                "bookName": name,
                "chapter": chapter,
                "verses": [{"verse": verse, "text": chapter_data[verse]} for verse in verse_numbers],
            }
            target = ASSET_DIR / book_id / f"{chapter}.json"
            target.parent.mkdir(parents=True, exist_ok=True)
            target.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
        books.append({
            "id": book_id,
            "name": name,
            "order": order,
            "aliases": [name, *aliases],
            "oneChapter": one_chapter,
            "verseCounts": verse_counts,
            "availableVerses": [sorted(book_data[chapter]) for chapter in chapters],
        })

    METADATA_PATH.write_text(
        json.dumps({"version": 1, "canon": "Protestant 66-book canon", "books": books}, ensure_ascii=False, separators=(",", ":")) + "\n",
        encoding="utf-8",
    )
    PROVENANCE_PATH.write_text(
        json.dumps({
            "version": 1,
            "translation": "Berean Standard Bible (BSB)",
            "license": "Public Domain",
            "source": "https://berean.bible/downloads.htm",
            "importSource": "https://bereanbible.com/bsb_usfm.zip",
            "sourceFile": source.name,
            "sourceSha256": source_sha,
            "assetFormat": "one deterministic JSON file per canonical chapter",
        }, ensure_ascii=False, separators=(",", ":")) + "\n",
        encoding="utf-8",
    )
    validate()


def validate() -> None:
    data = json.loads(METADATA_PATH.read_text(encoding="utf-8"))
    books = data.get("books") or []
    if len(books) != 66:
        raise ValueError(f"Expected 66 books, found {len(books)}")
    for order, (expected_id, expected_name, _aliases, one_chapter) in enumerate(CANON, start=1):
        book = books[order - 1]
        if book.get("id") != expected_id or book.get("name") != expected_name or book.get("order") != order:
            raise ValueError(f"Invalid canonical ordering at book {order}")
        counts = book.get("verseCounts")
        if not isinstance(counts, list) or not counts or not all(isinstance(count, int) and count > 0 for count in counts):
            raise ValueError(f"{expected_name} is missing valid chapter verse counts")
        if bool(book.get("oneChapter")) != one_chapter:
            raise ValueError(f"{expected_name} has incorrect one-chapter metadata")
        if one_chapter and len(counts) != 1:
            raise ValueError(f"{expected_name} must have exactly one chapter")
        for chapter in range(1, len(counts) + 1):
            asset = ASSET_DIR / expected_id / f"{chapter}.json"
            if not asset.is_file():
                raise ValueError(f"Missing BSB asset {asset}")
            payload = json.loads(asset.read_text(encoding="utf-8"))
            if payload.get("book") != expected_id or payload.get("chapter") != chapter:
                raise ValueError(f"Invalid BSB asset {asset}")
            actual = [item.get("verse") for item in payload.get("verses") or []]
            if not actual or max(actual) != counts[chapter - 1]:
                raise ValueError(f"Incorrect verse bounds in BSB asset {asset}")
            available = (book.get("availableVerses") or [])[chapter - 1]
            if actual != available:
                raise ValueError(f"Incorrect available verse list in BSB asset {asset}")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path, nargs="?")
    parser.add_argument("--check", action="store_true")
    args = parser.parse_args()
    if args.check:
        validate()
        print("BSB metadata and chapter assets are valid")
        return 0
    if not args.source or not args.source.is_file():
        parser.error("provide the downloaded bsb_tables.xlsx source, or use --check")
    build(args.source)
    print(f"Wrote BSB assets to {ASSET_DIR}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
